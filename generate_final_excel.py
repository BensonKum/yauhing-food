import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

# Read products.json
f = open('C:/Users/admin/.qclaw/workspace/yauhing-food/products.json', 'r', encoding='utf-8-sig')
inv_data = json.load(f)
f.close()

# Read website prices
f = open('C:/Users/admin/.qclaw/workspace/yauhing-food/website_prices.json', 'r', encoding='utf-8')
web_prices = json.load(f)
f.close()

# Read index.html to check which products exist on website
f = open('C:/Users/admin/.qclaw/workspace/yauhing-food/index.html', 'r', encoding='utf-8')
html = f.read()
f.close()

# Build data by category
cat_data = {}
for p in inv_data:
    cat = p.get('cat', '未分類')
    name = p.get('name', '')
    price = p.get('price', '')
    note = p.get('note', '')
    
    if cat not in cat_data:
        cat_data[cat] = []
    
    # Check if in website
    in_website = name in html
    
    # Get website price
    web_price = web_prices.get(name, '')
    
    cat_data[cat].append({
        'name': name,
        'inv_price': price,
        'inv_note': note,
        'in_website': in_website,
        'web_price': web_price
    })

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = "產品比對"

# Headers
headers = ['分類', 
           '倉存系統\n產品名稱', '倉存系統\n價錢', '倉存系統\n包裝規格',
           '祐興網站\n產品名稱', '祐興網站\n價錢', '祐興網站\n包裝規格',
           '備註']

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row = 2
for cat in sorted(cat_data.keys()):
    products = cat_data[cat]
    
    # Category header
    ws.cell(row=row, column=1, value=cat)
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    ws.cell(row=row, column=1).alignment = Alignment(vertical='center')
    
    ws.cell(row=row, column=2, value=f"共 {len(products)} 款")
    web_count = sum(1 for p in products if p['in_website'])
    ws.cell(row=row, column=5, value=f"共 {web_count} 款")
    ws.cell(row=row, column=8, value=f"僅倉存有: {len(products) - web_count} 款")
    row += 1
    
    for p in products:
        ws.cell(row=row, column=2, value=p['name'])
        ws.cell(row=row, column=3, value=p['inv_price'])
        ws.cell(row=row, column=4, value=p['inv_note'])
        
        if p['in_website']:
            ws.cell(row=row, column=5, value=p['name'])
            ws.cell(row=row, column=6, value=p['web_price'])
            ws.cell(row=row, column=8, value='✓ 兩系統均有')
        else:
            ws.cell(row=row, column=8, value='✗ 僅倉存系統有')
            ws.cell(row=row, column=8).font = Font(color='FF0000')
        
        row += 1
    
    row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 20

ws.row_dimensions[1].height = 40

# Save
output_path = 'C:/Users/admin/.qclaw/workspace/yauhing-food/祐興食品_產品比對.xlsx'
wb.save(output_path)
print(f'✓ Excel 已更新：{output_path}')
print('\n新增網站價格列（F 欄）')
