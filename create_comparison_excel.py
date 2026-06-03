import sys, json, re
sys.stdout.reconfigure(encoding='utf-8')

# Read products.json
f = open('C:/Users/admin/.qclaw/workspace/yauhing-food/products.json', 'r', encoding='utf-8-sig')
data = json.load(f)
f.close()

# Read index.html
f2 = open('C:/Users/admin/.qclaw/workspace/yauhing-food/index.html', 'r', encoding='utf-8')
html = f2.read()
f2.close()

# Build comparison data by category
cat_data = {}
for p in data:
    cat = p.get('cat', '未分類')
    name = p.get('name', '')
    found = name in html
    if cat not in cat_data:
        cat_data[cat] = {'inventory': [], 'website': []}
    if found:
        cat_data[cat]['website'].append(name)
    else:
        cat_data[cat]['inventory'].append(name)

# Create Excel file
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "產品比對"
    
    # Header
    ws['A1'] = "分類"
    ws['B1'] = "倉存系統 (products.json)"
    ws['C1'] = "祐興網站 (index.html)"
    ws['D1'] = "備註"
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    for col in ['A1', 'B1', 'C1', 'D1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 2
    for cat in sorted(cat_data.keys()):
        # Category header
        ws[f'A{row}'] = cat
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws[f'B{row}'] = f"共 {len(cat_data[cat]['inventory']) + len(cat_data[cat]['website'])} 款"
        ws[f'C{row}'] = f"共 {len(cat_data[cat]['website'])} 款"
        ws[f'D{row}'] = f"網站無: {len(cat_data[cat]['inventory'])} 款"
        row += 1
        
        # Products in both systems
        for name in cat_data[cat]['website']:
            ws[f'B{row}'] = name
            ws[f'C{row}'] = name
            ws[f'D{row}'] = "✓ 兩系統均有"
            row += 1
        
        # Products only in inventory system
        for name in cat_data[cat]['inventory']:
            ws[f'B{row}'] = name
            ws[f'C{row}'] = ""
            ws[f'D{row}'] = "✗ 僅倉存系統有"
            ws[f'D{row}'].font = Font(color='FF0000')
            row += 1
        
        row += 1  # Empty row between categories
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    
    # Save
    output_path = 'C:/Users/admin/.qclaw/workspace/yauhing-food/祐興食品_產品比對.xlsx'
    wb.save(output_path)
    print(f'✓ Excel 文件已創建：{output_path}')
    
except ImportError:
    print('openpyxl 未安裝，使用 CSV 格式輸出')
    # Fallback to CSV
    output_path = 'C:/Users/admin/.qclaw/workspace/yauhing-food/祐興食品_產品比對.csv'
    f = open(output_path, 'w', encoding='utf-8-sig')
    f.write('分類,倉存系統,祐興網站,備註\n')
    for cat in sorted(cat_data.keys()):
        f.write(f'\n[{cat}]\n')
        for name in cat_data[cat]['website']:
            f.write(f'{cat},"{name}","{name}","✓ 兩系統均有"\n')
        for name in cat_data[cat]['inventory']:
            f.write(f'{cat},"{name}","","✗ 僅倉存系統有"\n')
    f.close()
    print(f'✓ CSV 文件已創建：{output_path}')
