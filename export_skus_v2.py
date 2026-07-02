import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SRC = r'C:\Users\admin\.qclaw\workspace\yauhing-food\products.json'
DST = r'C:\Users\admin\.qclaw\workspace\yauhing-food\products_sku_list.xlsx'

with open(SRC, 'r', encoding='utf-8-sig') as f:
    products = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = '產品SKU列表'

# Styles
hdr_font = Font(bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['序號', '產品編號', '名稱', '規格', '價錢']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

row_idx = 2
for idx, p in enumerate(products):
    sku_val = p.get('sku', f'YH{str(idx+1).zfill(3)}')
    name = p.get('name', '')
    cat = p.get('cat', '')
    price = p.get('price', '')

    packs = price.split('|') if '|' in price else [price]
    packs = [p.strip() for p in packs if p.strip()]

    if not packs:
        packs = [price]

    for i, pack in enumerate(packs):
        alt = (row_idx - 2) % 2
        row_fill = PatternFill(
            start_color='EEF4FB' if alt else 'FFFFFF',
            end_color='EEF4FB' if alt else 'FFFFFF',
            fill_type='solid'
        )
        cells = [idx + 1, sku_val if i == 0 else '', name if i == 0 else '', cat, pack]
        for col, val in enumerate(cells, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(
                horizontal='center' if col in (1, 2, 5) else 'left',
                vertical='center'
            )
            cell.border = border
            cell.fill = row_fill
        row_idx += 1

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 32
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20

ws.freeze_panes = 'A2'
ws.row_dimensions[1].height = 22

wb.save(DST)
print(f'Done: {len(products)} products, {row_idx-2} rows, {os.path.getsize(DST)//1024}KB -> {DST}')
