# -*- coding: utf-8 -*-
"""
Merge noodlesx_all_products.json into the comparison Excel as Sheet2.
"""
import json
from openpyxl import load_workbook

# Load noodlesx products
with open('C:/Users/admin/.qclaw/workspace/yauhing-food/noodlesx_all_products.json', 'r', encoding='utf-8') as f:
    noodlesx = json.load(f)

print(f'noodlesx.com total: {len(noodlesx)} products')

# Load existing Excel
wb = load_workbook('C:/Users/admin/.qclaw/workspace/yauhing-food/祐興食品_產品比對.xlsx')

# Create Sheet2
if 'noodlesx.com' in wb.sheetnames:
    del wb['noodlesx.com']
ws2 = wb.create_sheet('noodlesx.com')

# Header
headers = ['序號', '產品名稱', '價格 (HKD)', '庫存狀態']
for col, h in enumerate(headers, 1):
    ws2.cell(row=1, column=col, value=h)

# Style header
from openpyxl.styles import Font, PatternFill
blue_font = Font(bold=True, color='FFFFFF', size=11)
blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
for col in range(1, len(headers) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.font = blue_font
    cell.fill = blue_fill

# Data
for i, p in enumerate(noodlesx):
    row = i + 2
    ws2.cell(row=row, column=1, value=i + 1)
    ws2.cell(row=row, column=2, value=p['name'])
    ws2.cell(row=row, column=3, value=f'HKD {p["price"]}')
    status = '缺貨中' if p.get('out_of_stock') else '有貨'
    cell = ws2.cell(row=row, column=4, value=status)
    if p.get('out_of_stock'):
        cell.font = Font(color='FF0000')

# Column widths
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 12

# Save
wb.save('C:/Users/admin/.qclaw/workspace/yauhing-food/祐興食品_產品比對.xlsx')
print(f'Sheet2 created: {len(noodlesx)} products')
print(f'Saved to 祐興食品_產品比對.xlsx')
