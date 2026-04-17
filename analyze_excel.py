import openpyxl, json

xlsx = r'C:\Users\admin\Documents\祐興庫存管理\祐興分店庫存管理系統.xlsx'
wb = openpyxl.load_workbook(xlsx, data_only=True)

out = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        rows.append([str(c) if c is not None else '' for c in row])
    out[sheet_name] = rows

with open(r'C:\Users\admin\.qclaw\workspace\yauhing-food\excel_detail.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print('OK')
