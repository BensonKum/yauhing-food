"""
祐興分店庫存管理系統升級
增加功能：
1. 儀表板自動庫存計算（SUMIFS公式）
2. 低庫存提醒條件格式
3. 每日/每月銷售報表
4. 兩店庫存對比
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from copy import copy
import datetime

SRC = r'C:\Users\admin\Documents\祐興庫存管理\祐興分店庫存管理系統.xlsx'
DST = r'C:\Users\admin\Documents\祐興庫存管理\祐興分店庫存管理系統_升級版.xlsx'

wb = openpyxl.load_workbook(SRC)

# ── 1. 庫存儀表板：加 SUMIFS 公式 ──────────────────────────────────
ws_dash = wb['庫存儀表板']

YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ORANGE = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Row 3 = header (SKU/產品名稱/分類/...)
# Data rows = 4 to 62
DATA_START = 4
DATA_END = 62

# Add SUMIFS formulas to 庫存儀表板
# Cols: D=中環初始(4), E=中環進貨(5), F=中環銷售(6), G=中環當前(7)
#        H=始創初始(8), I=始創進貨(9), J=始創銷售(10), K=始創當前(11)
#        L=總庫存(12), M=狀態(13), N=備註(14)
# 發票記錄: A=單據類型, B=單據編號, C=日期, D=產品名稱, E=分類, F=單位, G=數量, H=單價, I=分店, J=備註

for row in range(DATA_START, DATA_END + 1):
    sku_cell = ws_dash.cell(row=row, column=1).value
    if not sku_cell or not str(sku_cell).startswith('YH'):
        continue

    B = f'B{row}'  # 產品名稱
    F = f'F{row}'  # 分類

    # 中環進貨 (E): SUMIFS from 發票記錄 where 單據類型=進貨, 分店=中環, 產品名=B
    ws_dash.cell(row=row, column=5).value = (
        f'=IFERROR(SUMIFS(\'發票記錄\'!$G:$G,'
        f'\'發票記錄\'!$A:$A,"進貨",'
        f'\'發票記錄\'!$D:$D,{B},'
        f'\'發票記錄\'!$I:$I,"中環"),0)'
    )
    # 中環銷售 (F)
    ws_dash.cell(row=row, column=6).value = (
        f'=IFERROR(SUMIFS(\'發票記錄\'!$G:$G,'
        f'\'發票記錄\'!$A:$A,"銷售",'
        f'\'發票記錄\'!$D:$D,{B},'
        f'\'發票記錄\'!$I:$I,"中環"),0)'
    )
    # 中環當前 (G) = D + E - F
    ws_dash.cell(row=row, column=7).value = f'=D{row}+E{row}-F{row}'
    # 始創進貨 (I)
    ws_dash.cell(row=row, column=9).value = (
        f'=IFERROR(SUMIFS(\'發票記錄\'!$G:$G,'
        f'\'發票記錄\'!$A:$A,"進貨",'
        f'\'發票記錄\'!$D:$D,{B},'
        f'\'發票記錄\'!$I:$I,"始創"),0)'
    )
    # 始創銷售 (J)
    ws_dash.cell(row=row, column=10).value = (
        f'=IFERROR(SUMIFS(\'發票記錄\'!$G:$G,'
        f'\'發票記錄\'!$A:$A,"銷售",'
        f'\'發票記錄\'!$D:$D,{B},'
        f'\'發票記錄\'!$I:$I,"始創"),0)'
    )
    # 始創當前 (K) = H + I - J
    ws_dash.cell(row=row, column=11).value = f'=H{row}+I{row}-J{row}'
    # 總庫存 (L) = G + K
    ws_dash.cell(row=row, column=12).value = f'=G{row}+K{row}'
    # 狀態 (M): 顯示中文提示
    ws_dash.cell(row=row, column=13).value = (
        f'=IF(L{row}<=0,"缺貨",IF(L{row}<5,"庫存緊張",IF(L{row}<15,"偏低","庫存充足")))'
    )

    # 格式化
    for col in range(4, 15):
        c = ws_dash.cell(row=row, column=col)
        c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center')
        if col in [4, 8]:  # 初始庫存格 - 淺黃色，用戶可手動填
            c.fill = YELLOW
        elif col in [13]:  # 狀態格
            pass

# 條件格式：狀態顏色
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)
YELLOW_FILL2 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT2 = Font(color="9C6500", bold=True)

ws_dash.conditional_formatting.add(
    f'M{DATA_START}:M{DATA_END}',
    FormulaRule(formula=[f'L{DATA_START}<=0'], fill=RED_FILL, font=RED_FONT)
)
ws_dash.conditional_formatting.add(
    f'M{DATA_START}:M{DATA_END}',
    FormulaRule(formula=[f'AND(L{DATA_START}>0,L{DATA_START}<5)'], fill=YELLOW_FILL2, font=YELLOW_FONT2)
)

# 庫存數字條件格式
ws_dash.conditional_formatting.add(
    f'L{DATA_START}:L{DATA_END}',
    CellIsRule(operator='lessThan', formula=['5'], fill=RED_FILL, font=RED_FONT)
)

# 標題行格式
for col in range(1, 15):
    c = ws_dash.cell(row=3, column=col)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN

# 調整列寬
col_widths = [8, 22, 10, 8, 8, 8, 8, 8, 8, 8, 8, 9, 12, 15]
for i, w in enumerate(col_widths, 1):
    ws_dash.column_dimensions[get_column_letter(i)].width = w

ws_dash.row_dimensions[3].height = 30

print("[OK] 儀表板公式已加")

# ── 2. 新增：庫存提醒工作表 ──────────────────────────────────────
if '庫存提醒' in wb.sheetnames:
    del wb['庫存提醒']

ws_alert = wb.create_sheet('庫存提醒')
ws_alert.column_dimensions['A'].width = 8
ws_alert.column_dimensions['B'].width = 22
ws_alert.column_dimensions['C'].width = 10
ws_alert.column_dimensions['D'].width = 12
ws_alert.column_dimensions['E'].width = 12
ws_alert.column_dimensions['F'].width = 12
ws_alert.column_dimensions['G'].width = 12
ws_alert.column_dimensions['H'].width = 10
ws_alert.column_dimensions['I'].width = 15

# Header
alerts_header = ['SKU', '產品名稱', '分類', '中環當前', '始創當前', '總庫存', '狀態', '備註']
for col, h in enumerate(alerts_header, 1):
    c = ws_alert.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = THIN

# Formula rows linking to dashboard
for i, row in enumerate(range(DATA_START, DATA_END + 1)):
    r = i + 2
    sku = ws_dash.cell(row=row, column=1).value
    if not sku or not str(sku).startswith('YH'):
        continue
    dash_row = row
    ws_alert.cell(row=r, column=1, value=f'=庫存儀表板!A{dash_row}')
    ws_alert.cell(row=r, column=2, value=f'=庫存儀表板!B{dash_row}')
    ws_alert.cell(row=r, column=3, value=f'=庫存儀表板!C{dash_row}')
    ws_alert.cell(row=r, column=4, value=f'=庫存儀表板!G{dash_row}')
    ws_alert.cell(row=r, column=5, value=f'=庫存儀表板!K{dash_row}')
    ws_alert.cell(row=r, column=6, value=f'=庫存儀表板!L{dash_row}')
    ws_alert.cell(row=r, column=7, value=f'=庫存儀表板!M{dash_row}')
    for col in range(1, 9):
        ws_alert.cell(row=r, column=col).border = THIN
        ws_alert.cell(row=r, column=col).alignment = Alignment(horizontal='center')

# 條件格式：狀態
ws_alert.conditional_formatting.add(
    f'A2:H{60}',
    FormulaRule(formula=['$G2="缺貨"'], fill=RED_FILL, font=RED_FONT)
)
ws_alert.conditional_formatting.add(
    f'A2:H{60}',
    FormulaRule(formula=['$G2="庫存緊張"'], fill=YELLOW_FILL2, font=YELLOW_FONT2)
)

print("[OK] 庫存提醒工作表已加")

# ── 3. 新增：銷售報表 ────────────────────────────────────────────
if '銷售報表' in wb.sheetnames:
    del wb['銷售報表']

ws_rep = wb.create_sheet('銷售報表')
ws_rep.column_dimensions['A'].width = 10
ws_rep.column_dimensions['B'].width = 10
ws_rep.column_dimensions['C'].width = 15
ws_rep.column_dimensions['D'].width = 22
ws_rep.column_dimensions['E'].width = 10
ws_rep.column_dimensions['F'].width = 10
ws_rep.column_dimensions['G'].width = 10
ws_rep.column_dimensions['H'].width = 10
ws_rep.column_dimensions['I'].width = 10
ws_rep.column_dimensions['J'].width = 15

today = datetime.date.today().strftime('%Y/%m/%d')

# Section 1: 每日銷售
ws_rep.cell(row=1, column=1, value='祐興食品 — 銷售報表')
ws_rep.cell(row=1, column=1).font = Font(bold=True, size=14)

ws_rep.cell(row=2, column=1, value=f'製表日期：{today}')

ws_rep.cell(row=4, column=1, value='【每日銷售】')
ws_rep.cell(row=4, column=1).font = Font(bold=True, size=12)

daily_header = ['日期', '分店', '銷售筆數', '銷售額(HKD)']
for col, h in enumerate(daily_header, 1):
    c = ws_rep.cell(row=5, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center')
    c.border = THIN

# Extract unique dates from 發票記錄
ws_rec = wb['發票記錄']
dates = set()
for row in ws_rec.iter_rows(min_row=4, values_only=True):
    if row[0] == '銷售' and row[2]:
        dates.add(str(row[2])[:10])

sorted_dates = sorted(dates)
for i, d in enumerate(sorted_dates):
    r = 6 + i
    ws_rep.cell(row=r, column=1, value=d).border = THIN
    ws_rep.cell(row=r, column=2, value='中環').border = THIN
    ws_rep.cell(row=r, column=3,
        value=f'=IFERROR(COUNTIFS(\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$C:$C,"{d}",\'發票記錄\'!$I:$I,"中環"),0)'
    ).border = THIN
    ws_rep.cell(row=r, column=4,
        value=f'=IFERROR(SUMIFS(\'發票記錄\'!$H:$H,\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$C:$C,"{d}",\'發票記錄\'!$I:$I,"中環")*COUNTIFS(\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$C:$C,"{d}",\'發票記錄\'!$I:$I,"中環"),0)'
    ).border = THIN
    for col in range(1, 5):
        ws_rep.cell(row=r, column=col).alignment = Alignment(horizontal='center')

    ws_rep.cell(row=r+50, column=1, value=d).border = THIN
    ws_rep.cell(row=r+50, column=2, value='始創').border = THIN
    ws_rep.cell(row=r+50, column=3,
        value=f'=IFERROR(COUNTIFS(\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$C:$C,"{d}",\'發票記錄\'!$I:$I,"始創"),0)'
    ).border = THIN
    ws_rep.cell(row=r+50, column=4,
        value=f'=IFERROR(SUMIFS(\'發票記錄\'!$H:$H,\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$C:$C,"{d}",\'發票記錄\'!$I:$I,"始創")*COUNTIFS(\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$C:$C,"{d}",\'發票記錄\'!$I:$I,"始創"),0)'
    ).border = THIN
    for col in range(1, 5):
        ws_rep.cell(row=r+50, column=col).alignment = Alignment(horizontal='center')

# Section 2: 按產品銷售
prod_row = 60 + len(sorted_dates) * 2 + 5
ws_rep.cell(row=prod_row, column=1, value='【按產品銷售統計】')
ws_rep.cell(row=prod_row, column=1).font = Font(bold=True, size=12)

prod_header = ['產品名稱', '分類', '中環銷量', '始創銷量', '總銷量', '中環銷售額', '始創銷售額', '總銷售額']
for col, h in enumerate(prod_header, 1):
    c = ws_rep.cell(row=prod_row + 1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center')
    c.border = THIN

ws_prod = wb['產品列表']
prod_count = 0
for i, row in enumerate(range(4, 67)):
    sku = ws_prod.cell(row=row, column=1).value
    name = ws_prod.cell(row=row, column=2).value
    cat = ws_prod.cell(row=row, column=3).value
    price_val = ws_prod.cell(row=row, column=4).value
    if not sku or not str(sku).startswith('YH'):
        continue
    r = prod_row + 2 + prod_count
    ws_rep.cell(row=r, column=1, value=name).border = THIN
    ws_rep.cell(row=r, column=2, value=cat).border = THIN
    ws_rep.cell(row=r, column=3,
        value=f'=IFERROR(SUMIFS(\'發票記錄\'!$G:$G,\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$D:$D,A{r},\'發票記錄\'!$I:$I,"中環"),0)'
    ).border = THIN
    ws_rep.cell(row=r, column=4,
        value=f'=IFERROR(SUMIFS(\'發票記錄\'!$G:$G,\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$D:$D,A{r},\'發票記錄\'!$I:$I,"始創"),0)'
    ).border = THIN
    ws_rep.cell(row=r, column=5, value=f'=C{r}+D{r}').border = THIN
    price_ref = f'IFERROR(VLOOKUP(A{r},\'產品列表\'!$B:$D,3,FALSE),{price_val if price_val else 0})'
    ws_rep.cell(row=r, column=6, value=f'=C{r}*{price_ref}').border = THIN
    ws_rep.cell(row=r, column=7, value=f'=D{r}*{price_ref}').border = THIN
    ws_rep.cell(row=r, column=8, value=f'=F{r}+G{r}').border = THIN
    for col in range(1, 9):
        ws_rep.cell(row=r, column=col).alignment = Alignment(horizontal='center')
    prod_count += 1

# Section 3: 按月統計
month_row = prod_row + 2 + prod_count + 3
ws_rep.cell(row=month_row, column=1, value='【按月統計】')
ws_rep.cell(row=month_row, column=1).font = Font(bold=True, size=12)

month_header = ['月份', '中環銷售筆數', '始創銷售筆數', '中環銷售額(HKD)', '始創銷售額(HKD)', '總銷售額(HKD)']
for col, h in enumerate(month_header, 1):
    c = ws_rep.cell(row=month_row + 1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center')
    c.border = THIN

# Get unique months
months = set()
for row in ws_rec.iter_rows(min_row=4, values_only=True):
    if row[0] == '銷售' and row[2]:
        months.add(str(row[2])[:7])

sorted_months = sorted(months)
for i, m in enumerate(sorted_months):
    r = month_row + 2 + i
    ws_rep.cell(row=r, column=1, value=m).border = THIN
    ws_rep.cell(row=r, column=2,
        value=f'=IFERROR(COUNTIFS(\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$C:$C,">="&"{m}/01",\'發票記錄\'!$C:$C,"<"&TEXT(DATEVALUE("{m}/01")+30,"YYYY/MM/DD"),\'發票記錄\'!$I:$I,"中環"),0)'
    ).border = THIN
    ws_rep.cell(row=r, column=3,
        value=f'=IFERROR(COUNTIFS(\'發票記錄\'!$A:$A,"銷售",\'發票記錄\'!$C:$C,">="&"{m}/01",\'發票記錄\'!$C:$C,"<"&TEXT(DATEVALUE("{m}/01")+30,"YYYY/MM/DD"),\'發票記錄\'!$I:$I,"始創"),0)'
    ).border = THIN
    ws_rep.cell(row=r, column=4,
        value=f'=IFERROR(SUMPRODUCT(\'發票記錄\'!$H:$H*(\'發票記錄\'!$A:$A="銷售")*(\'發票記錄\'!$I:$I="中環")*(LEFT(\'發票記錄\'!$C:$C,7)="{m}")),0)'
    ).border = THIN
    ws_rep.cell(row=r, column=5,
        value=f'=IFERROR(SUMPRODUCT(\'發票記錄\'!$H:$H*(\'發票記錄\'!$A:$A="銷售")*(\'發票記錄\'!$I:$I="始創")*(LEFT(\'發票記錄\'!$C:$C,7)="{m}")),0)'
    ).border = THIN
    ws_rep.cell(row=r, column=6, value=f'=D{r}+E{r}').border = THIN
    for col in range(1, 7):
        ws_rep.cell(row=r, column=col).alignment = Alignment(horizontal='center')

print("[OK] 銷售報表已加")

# ── 4. 發票記錄：加 AutoFilter + 格式 ────────────────────────────
ws_rec = wb['發票記錄']
for col in range(1, 11):
    c = ws_rec.cell(row=3, column=col)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.border = THIN
    c.alignment = Alignment(horizontal='center')
ws_rec.auto_filter.ref = f"A3:J{ws_rec.max_row}"
print("[OK] 發票記錄 AutoFilter 已加")

# ── 5. 儀表板加凍結窗格 ─────────────────────────────────────────
ws_dash.freeze_panes = 'D4'
print("[OK] 儀表板凍結窗格已設")

# ── 6. 產品列表加 AutoFilter ─────────────────────────────────────
ws_prod = wb['產品列表']
ws_prod.auto_filter.ref = f"A2:F{ws_prod.max_row}"
ws_prod.cell(row=2, column=1).fill = HEADER_FILL
ws_prod.cell(row=2, column=1).font = HEADER_FONT
print("[OK] 產品列表 AutoFilter 已加")

# Save
wb.save(DST)
print(f"\n[OK] 升級完成！")
print(f"   保存至：{DST}")
print(f"   新增工作表：庫存提醒、銷售報表")
print(f"   更新工作表：庫存儀表板（公式）、發票記錄（AutoFilter）、產品列表（AutoFilter）")
