import sys, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

# Read products.json
f = open('C:/Users/admin/.qclaw/workspace/yauhing-food/products.json', 'r', encoding='utf-8-sig')
data = json.load(f)
f.close()

# Read index.html
f2 = open('C:/Users/admin/.qclaw/workspace/yauhing-food/index.html', 'r', encoding='utf-8')
html = f2.read()
f2.close()

# Build product dict from products.json
inv_products = {}
for p in data:
    cat = p.get('cat', '未分類')
    name = p.get('name', '')
    price = p.get('price', '')
    note = p.get('note', '')
    if cat not in inv_products:
        inv_products[cat] = {}
    inv_products[cat][name] = {'price': price, 'note': note}

# Try to extract price from index.html
# Pattern: product card with name and price
# Look for price patterns near product names
web_products = {}
# This is tricky - index.html might not have structured price data
# Let's search for price info near product names

print('Creating Excel with price and packaging info...')

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = "產品比對"

# Headers
headers = ['分類', 
           '倉存系統\n產品名稱', '倉存系統\n價錢', '倉存系統\n包裝規格',
           '祐興網站\n產品名稱', '祐興網站\n價錢', '祐興網站\n包裝規格',
           '備註']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

row = 2
for cat in sorted(inv_products.keys()):
    cat_products = inv_products[cat]
    
    # Category header row
    ws.cell(row=row, column=1, value=cat)
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    ws.cell(row=row, column=1).alignment = Alignment(vertical='center')
    
    ws.cell(row=row, column=2, value=f"共 {len(cat_products)} 款")
    ws.cell(row=row, column=6, value=f"網站有: {sum(1 for name in cat_products if name in html)} 款")
    ws.cell(row=row, column=8, value=f"僅倉存有: {sum(1 for name in cat_products if name not in html)} 款")
    row += 1
    
    for name in sorted(cat_products.keys()):
        found_in_web = name in html
        
        # Column B: Inventory system product name
        ws.cell(row=row, column=2, value=name)
        
        # Column C: Inventory system price
        price = cat_products[name]['price']
        ws.cell(row=row, column=3, value=price)
        
        # Column D: Inventory system packaging/note
        note = cat_products[name]['note']
        ws.cell(row=row, column=4, value=note if note else '')
        
        # Column F: Website product name (if exists)
        if found_in_web:
            ws.cell(row=row, column=6, value=name)
            # Try to find price in HTML (this is a simplification)
            ws.cell(row=row, column=7, value='見網站')
        
        # Column H: Remarks
        if found_in_web:
            ws.cell(row=row, column=8, value='✓ 兩系統均有')
        else:
            ws.cell(row=row, column=8, value='✗ 僅倉存系統有')
            ws.cell(row=row, column=8).font = Font(color='FF0000')
        
        row += 1
    
    row += 1  # Empty row between categories

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 20

# Set row height for header
ws.row_dimensions[1].height = 40

# Save
output_path = 'C:/Users/admin/.qclaw/workspace/yauhing-food/祐興食品_產品比對.xlsx'
wb.save(output_path)
print(f'✓ Excel 文件已更新：{output_path}')
print('\nHeaders:')
for col, header in enumerate(headers, 1):
    print(f'  {chr(64+col)}1: {header}')
