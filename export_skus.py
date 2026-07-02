import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\admin\.qclaw\workspace\yauhing-food\products.json'
DST = r'C:\Users\admin\.qclaw\workspace\yauhing-food\products_sku_list.xlsx'

with open(SRC, 'r', encoding='utf-8-sig') as f:
    products = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = '產品SKU列表'

# Header style
hdr_font = Font(bold=True, color='FFFFFF', size=12)
hdr_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['序號', 'SKU', '名稱', '規格', '價錢']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

row_idx = 2
for idx, p in enumerate(products):
    sku_val = f'YH{str(idx+1).zfill(3)}'
    name = p.get('name', '')
    cat = p.get('cat', '')
    price = p.get('price', '')

    # Multi-pack: split into lines
    packs = []
    if '|' in price:
        parts = price.split('|')
        for part in parts:
            part = part.strip()
            if part:
                packs.append(part)
    else:
        packs = [price]

    for i, pack in enumerate(packs):
        sku_disp = sku_val if i == 0 else ''
        alt = (i % 2)
        row_fill = PatternFill(start_color='EEF4FB' if alt else 'FFFFFF',
                               end_color='EEF4FB' if alt else 'FFFFFF',
                               fill_type='solid')
        cells = ['', sku_disp, name if i == 0 else '', cat, pack]
        for col, val in enumerate(cells, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal='center' if col in (1,2,5) else 'left',
                                       vertical='center')
            cell.border = border
            cell.fill = row_fill
        row_idx += 1

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 18

# Freeze header
ws.freeze_panes = 'A2'

wb.save(DST)
print(f'Done: {row_idx-2} rows, {os.path.getsize(DST)//1024}KB -> {DST}')
